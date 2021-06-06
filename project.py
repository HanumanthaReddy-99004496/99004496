
"""
project.py
program to show all ps.no and all columns
print user wanted column to new excel file
"""
# pylint: disable=R1705
import openpyxl

wb = openpyxl.load_workbook("input.xlsx")
all_sheets = wb.sheetnames
# view existing sheets
validate_psNumbers = []
titles = []
fetched_data = []


def view_psnumbers():
    """function to display ps.numbers"""
    sheet = wb["academic"]  # selecting particular sheet.
    rows = sheet.max_row  # number of columns in a sheet including heading row
    for j in range(1, rows + 1):
        # print all cells under column1 of every sheet which is ps number
        psno = sheet.cell(row=j, column=1)
        validate_psNumbers.append(psno.value)
    return validate_psNumbers


def writeto_outputxl():
    """ write output to new xl file"""
    try:
        out = openpyxl.Workbook()  # create a new blank workbook
        sheet = out.active
        sheet.title = "Output"
        for no_cols in range(1, 21):
            row_1 = sheet.cell(row=1, column=no_cols)
            row_1.value = titles[no_cols - 1]
        for no_cols in range(1, 21):
            row_2 = sheet.cell(row=2, column=no_cols)
            row_2.value = fetched_data[no_cols - 1]
        out.save("output.xlsx")
        return "Written output to 'output.xlsx'"
    except PermissionError:
        return "Errors writing please!, close 'output.xl' file if opened"


class Excel:
    """
    class contains methods to validate ps numbers and column entered by user
    and get data that user need using get_data method
    """

    def __init__(self):
        self.sheet_name = ""
        self.number = ""

    def validate(self, number):
        """validate user entered ps number"""
        self.number = number
        if self.number in validate_psNumbers:
            return "Valid"
        else:
            return "Invalid PS.No"

    def validate_sheet(self, sheet_name):
        """function to validate entered column name"""
        self.sheet_name = sheet_name
        if self.sheet_name in all_sheets:
            return "Valid"
        else:
            return "Invalid"

    def get_data(self, psnumber, sheet_name):
        """get requested data"""
        self.number = psnumber
        self.sheet_name = sheet_name
        data = wb[sheet_name]
        row_id = validate_psNumbers.index(psnumber) + 1
        columns = data.max_column
        for col_val in range(1, columns + 1):
            fetch_titles = data.cell(row=1, column=col_val)
            # print(fetch_titles.value, end="\t\t")   # column names
            # storing column names in list to write it to output.xlsx
            titles.append(fetch_titles.value)
        print()
        for col_val in range(1, columns + 1):
            fetch_data = data.cell(row=row_id, column=col_val)
            # print(fetch_data.value, end="\t\t")    # requested data
            # storing  data in list to write it to output.xlsx
            fetched_data.append(fetch_data.value)
        return fetched_data


if __name__ == "__main__":

    obj = Excel()  # creating class object
    view = view_psnumbers()
    for i in view:
        print(i)
    ps = int(input("Enter ps Number from above list: "))
    PS_CHECK = obj.validate(ps)
    if PS_CHECK == "Valid":
        print("\nData in excel")
        for i in all_sheets:
            print(i + "\t", end="")
        input_sheet = input("\n\nEnter data/sheet name from above list: ")
        SHEET_CHECK = obj.validate_sheet(input_sheet)
        if SHEET_CHECK == "Invalid":
            print("Invalid Data Requested")
        else:
            obj.get_data(ps, input_sheet)
            VAL1 = writeto_outputxl()
            print(VAL1)
    else:
        print(PS_CHECK)
